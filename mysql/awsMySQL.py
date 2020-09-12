# encoding = UTF-8

"""
    Oracle DB Handling...

    1. Select
    2. Insert
    3. Update
    4. Delete
"""
import pymysql
# Excel 처
from openpyxl import Workbook
from openpyxl import load_workbook

def connectDB():
    dbname = "mercury"
    charset = "utf8"

    connDB = pymysql.connect(host=url, user=userid, password=passwd,
                             db=dbname, charset=charset)
    return connDB


def selectItem():
    conn = connectDB()
    sql = "select * from user"

    # 모든 함수들은 connection Leak를 막아주기 위해 try ..finally 구문을 사용하고,
    # with문으로 cursor리소스를 자동으로 해제되도록 코드를 작성한다.
    try:
        # with conn.cursor() as curs:
        with conn.cursor(pymysql.cursors.DictCursor) as curs:
            curs.execute(sql)
            rs = curs.fetchall()
            for row in rs:
                print(row)
    finally:
        conn.close()


def insertItem():
    conn = connectDB()
    try:
        with conn.cursor() as curs:
            sql = 'insert into user (USER_ID, USER_ACCOUNT) values(%s, %s)'
            # execute() 계열 메서드 첫번째 파라미터에는 SQL을, 두번째 파라미터에 실제 데이터를 리스트나 튜플 형태로
            # 문자열, 숫자 등에 관계 없이 대치할 값은 모두 %s로 쓰입니다
            data = (22, "phg4u")
            curs.execute(sql, data)
            conn.commit()
    finally:
        conn.close()


def updateItem():
    conn = connectDB()
    try:
        with conn.cursor() as curs:
            sql = 'update user set email=%s where user_id=%s'
            curs.execute(sql, ("parhg@gmail.com", 1))
            # curs.execute(sql, (test_obj.name, test_obj.num))

            conn.commit()
    finally:
        conn.close()


def deleteItem():
    conn = connectDB()
    try:
        with conn.cursor() as curs:
            sql = 'delete from user where user_id=%s'
            curs.execute(sql, 22)
            conn.commit()
    finally:
        conn.close()

# create table test(num int(11), name varchar(10));
def insertExcelFile():
    conn = connectDB()
    try:
        with conn.cursor() as curs:
            sql = 'insert into test values(%s, %s)'
            wb = load_workbook('숫자.xlsx', data_only=True)
            ws = wb['Sheet']

            # iter를 사용한 이유는 엑셀파일의 첫번째 행을 빼고 반복문을 돌릴려고 사용했다.
            iter_rows = iter(ws.rows)
            next(iter_rows)
            for row in iter_rows:
                curs.execute(sql, (row[0].value, row[1].value))
            conn.commit()
    finally:
        conn.close()


def SelectData2Excel():

    conn = connectDB()
    try:
        with conn.cursor() as curs:
            sql = "select * from test"
            curs.execute(sql)
            rs = curs.fetchall()

            wb = Workbook()
            ws = wb.active

            # 첫행 입력
            ws.append(('번호', '이름'))

            # DB 모든 데이터 엑셀로
            for row in rs:
                ws.append(row)

            wb.save('숫자.xlsx')
    finally:
        conn.close()
        wb.close()


def run():
    # insertItem()
    # updateItem()
    # deleteItem()
    selectItem()


if __name__ == '__main__':
    run()
